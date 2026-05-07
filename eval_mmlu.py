"""MMLU-Pro evaluation using ROMA's executor.

Usage:
    uv run python eval_mmlu.py
    uv run python eval_mmlu.py --model deepseek-v4-flash --limit 50
"""

from __future__ import annotations

import argparse
import logging
import os
import random
import re
import sys
from pathlib import Path

import datasets
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

load_dotenv()

from src.components import build_default_registry
from src.config.loader import load_config
from src.core.models import Task
from src.core.signatures import ExecutorOutput

random.seed(42)
_log = logging.getLogger("eval_mmlu")

CATEGORIES = [
    "computer science", "math", "chemistry", "engineering", "law", "biology",
    "health", "physics", "business", "philosophy", "economics", "other",
    "psychology", "history",
]


def form_options(options: list[str]) -> str:
    lines = ["Options are:"]
    opts = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J"]
    for opt, o in zip(options, opts):
        lines.append(f"({o}): {opt}")
    return "\n".join(lines)


def get_prediction(output: str) -> str:
    pattern = r"answer is \(?([ABCDEFGHIJ])\)?"
    match = re.search(pattern, output)
    if match:
        return match.group(1)
    _log.warning("extraction failed for output, random guess: %s", output[:80])
    return random.choice(["A", "B", "C", "D", "E", "F", "G", "H", "I", "J"])


def run_roma(
    query: str,
    model: str | None = None,
    api_keys_path: Path | None = None,
) -> tuple[str, list[str]]:
    """Run a single question through ROMA.

    Returns (final_output, intermediate_steps) where intermediate_steps
    captures subtask goals and their results from the event stream.
    """
    config = load_config()
    runtime = config.runtime

    if api_keys_path:
        _load_toml_keys(api_keys_path)
    _load_toml_keys(runtime.api_keys_path)

    registry = build_default_registry(
        tavily_api_key=_resolve("TAVILY_API_KEY", "tavily_api_key"),
        python_executable=sys.executable,
        limits=runtime.limits,
        verbose=False,
    )
    from src.core.controller import RomaController

    steps: list[str] = []

    def collector(kind: str, payload: dict, trace) -> None:
        task_id = trace.task_id if trace else "?"
        if kind == "task_started":
            goal = payload.get("goal", "") or (trace.goal if trace else "")
            steps.append(f"[start] {task_id}: {goal[:120]}")
        elif kind == "atomizer_decision":
            node = payload.get("node_type", "?")
            tools = payload.get("granted_tools", [])
            steps.append(f"[atomizer] {task_id}: {node}" + (f" tools={tools}" if tools else ""))
        elif kind == "planner_output":
            deps = payload.get("dependency_batches", [])
            flat = [tid for batch in deps for tid in batch]
            steps.append(f"[planner] {task_id}: subtasks={flat}")
        elif kind == "executor_completed":
            preview = (payload.get("result_preview", "") or "")[:200]
            steps.append(f"[executor] {task_id}: {preview}")
        elif kind == "executor_failed":
            err = payload.get("error", "?")
            steps.append(f"[executor] {task_id}: ERROR {err}")
        elif kind == "aggregation_completed":
            preview = (payload.get("summary_preview", "") or "")[:200]
            steps.append(f"[aggregator] {task_id}: {preview}")

    controller = RomaController(registry, event_callback=collector)
    root = Task(id="root", goal=query, context_input=None)
    outcome = controller.solve(root)
    return outcome.output.result, steps


def run_direct(
    query: str,
    model: str | None = None,
) -> str:
    """Run a single question through the LLM directly (bypass ROMA pipeline)."""
    from src.core.inference import build_client, get_default_model

    client = build_client()
    if client is None:
        return "Error: no LLM client available"

    chosen = model or get_default_model()
    messages = [
        {
            "role": "system",
            "content": (
                "You are a knowledge expert. Answer the multi-choice question "
                "and state your final answer as `The answer is ...`."
            ),
        },
        {"role": "user", "content": query},
    ]

    response = client.chat.completions.create(
        model=chosen,
        messages=messages,
        temperature=0.1,
        max_tokens=4096,
    )
    return response.choices[0].message.content or ""


def _resolve(*names: str) -> str | None:
    for n in names:
        v = os.getenv(n)
        if v:
            return v
    return None


def _load_toml_keys(path: Path) -> None:
    if not path.exists():
        return
    try:
        import tomllib

        data = tomllib.loads(path.read_text(encoding="utf-8"))
        for k, v in data.get("providers", {}).items():
            if isinstance(v, str) and v.strip() and not os.getenv(k.upper()):
                os.environ[k.upper()] = v.strip()
    except ImportError:
        pass


def build_workbook(results: list[dict]) -> Workbook:
    wb = Workbook()

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    def style_header(ws, ncols: int) -> None:
        for col in range(1, ncols + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

    # --- Summary sheet ---
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum.append(["Category", "Correct", "Total", "Accuracy"])
    style_header(ws_sum, 4)

    per_cat: dict[str, list[int]] = {c: [0, 0] for c in CATEGORIES}
    for r in results:
        cat = r["category"]
        if cat in per_cat:
            if r["correct"]:
                per_cat[cat][0] += 1
            else:
                per_cat[cat][1] += 1

    total_correct = 0
    total_all = 0
    for cat in CATEGORIES:
        c, f = per_cat[cat]
        total = c + f
        acc = c / total if total > 0 else 0.0
        ws_sum.append([cat, c, total, round(acc, 4)])
        total_correct += c
        total_all += total
        if total > 0:
            row = ws_sum.max_row
            fill = green_fill if acc >= 0.5 else red_fill
            for col in range(1, 5):
                ws_sum.cell(row=row, column=col).fill = fill

    ws_sum.append([])
    ws_sum.append(
        ["Overall", total_correct, total_all, round(total_correct / total_all, 4) if total_all > 0 else 0.0]
    )
    for col in range(1, 5):
        cell = ws_sum.cell(row=ws_sum.max_row, column=col)
        cell.font = Font(bold=True)

    for col in range(1, 5):
        ws_sum.column_dimensions[get_column_letter(col)].width = 20

    # --- Details sheet ---
    headers = ["#", "Category", "Question", "Correct Answer", "Prediction", "Correct", "Raw Output", "Intermediate Steps"]
    ws_det = wb.create_sheet("Details")
    ws_det.append(headers)
    style_header(ws_det, len(headers))

    for i, r in enumerate(results, 1):
        row = ws_det.max_row + 1
        steps_text = "\n".join(r.get("intermediate_steps", [])) if r.get("intermediate_steps") else ""
        ws_det.append([
            i,
            r["category"],
            r["question"],
            r["correct_answer"],
            r["prediction"],
            "Yes" if r["correct"] else "No",
            r["raw_output"],
            steps_text,
        ])
        fill = green_fill if r["correct"] else red_fill
        for col in range(1, len(headers) + 1):
            ws_det.cell(row=row, column=col).fill = fill

    ws_det.column_dimensions["A"].width = 6
    ws_det.column_dimensions["B"].width = 18
    ws_det.column_dimensions["C"].width = 60
    ws_det.column_dimensions["D"].width = 14
    ws_det.column_dimensions["E"].width = 14
    ws_det.column_dimensions["F"].width = 10
    ws_det.column_dimensions["G"].width = 100
    ws_det.column_dimensions["H"].width = 120

    return wb


def main() -> None:
    parser = argparse.ArgumentParser(description="MMLU-Pro evaluation with ROMA")
    parser.add_argument("--model", default=None, help="Override the LLM model")
    parser.add_argument("--limit", type=int, default=None, help="Limit number of test samples per category")
    parser.add_argument("--output", default="mmlu_results.xlsx", help="Output Excel file path")
    parser.add_argument("--direct", action="store_true", help="Use direct LLM call instead of ROMA pipeline")
    args = parser.parse_args()

    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")

    print("Loading MMLU-Pro dataset...")
    dataset = datasets.load_dataset("TIGER-Lab/MMLU-Pro")

    results: list[dict] = []

    test_data = list(dataset["test"])
    if args.limit:
        random.shuffle(test_data)
        test_data = test_data[: args.limit]

    total = len(test_data)
    print(f"Evaluating {total} questions...")

    for idx, entry in enumerate(test_data):
        cat = entry["category"]
        query = (
            "Q: "
            + entry["question"]
            + "\n"
            + "You must answer in this format: 'The answer is <Option>. '"
            + "\n"
            + form_options(entry["options"])
            + "\n"
        )

        if args.direct:
            answer_text = run_direct(query, model=args.model)
            steps: list[str] = []
        else:
            answer_text, steps = run_roma(query, model=args.model)

        prediction = get_prediction(answer_text)
        correct_answer = entry["answer"]
        is_correct = prediction == correct_answer

        results.append({
            "category": cat,
            "question": entry["question"],
            "correct_answer": correct_answer,
            "prediction": prediction,
            "correct": is_correct,
            "raw_output": answer_text,
            "intermediate_steps": steps,
        })

        acc_sofar = sum(1 for r in results if r["correct"]) / len(results)
        print(f"  [{idx + 1}/{total}] {cat:20s} pred={prediction} correct={correct_answer} {'✓' if is_correct else '✗'}  acc={acc_sofar:.4f}")

    wb = build_workbook(results)
    wb.save(args.output)
    print(f"\nResults saved to {args.output}")

    final_acc = sum(1 for r in results if r["correct"]) / len(results)
    print(f"Final accuracy: {final_acc:.4f} ({sum(1 for r in results if r['correct'])}/{len(results)})")


if __name__ == "__main__":
    main()
